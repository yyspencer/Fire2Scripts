#!/usr/bin/env python3
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SOURCE_XLSX = "Fire 2 Data.xlsx"
OUTPUT_XLSX = "Fire 2 TMP.xlsx"

# Solid yellow fill (works reliably in Excel)
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")

def is_four(val):
    """Return True if the cell value represents the number 4."""
    if val is None:
        return False
    # numeric 4 (int or float like 4.0)
    if isinstance(val, (int, float)):
        return int(val) == 4 and (val == int(val))
    # text '4'
    try:
        return str(val).strip() == "4"
    except Exception:
        return False

def main():
    # Duplicate the source workbook
    shutil.copyfile(SOURCE_XLSX, OUTPUT_XLSX)

    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.worksheets[0]  # first sheet

    highlighted = 0
    total = 0

    # Header row at 1 → start from row 2
    for row in range(2, ws.max_row + 1):
        total += 1
        cond_cell = ws.cell(row=row, column=2)  # column B (condition)
        if is_four(cond_cell.value):
            idx_cell = ws.cell(row=row, column=1)  # column A (index)
            idx_cell.fill = YELLOW_FILL
            highlighted += 1

    wb.save(OUTPUT_XLSX)
    print(f"Done. Checked {total} rows (starting at row 2). Highlighted {highlighted} index cell(s) in yellow.")
    print(f"Output written to: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()